import json
import os
import random

import openpyxl
import requests

def open_excel():#读取ececl数据
    wb = openpyxl.load_workbook(str(os.getcwd()+"/HelloWorld/ICBU/icbu_draft_orders.xlsx"))#TH泰国参数
    ws = wb.active
    data = []
    for i in range(99,100):#ws.max_row
        x=str(ws['A'+str(i)].value)
        data.append(json.loads(x))
    return data


def Icbu_Create():
    datas = open_excel()
    url = "http://120.24.31.239:20000/tms-saas-web/order/booking"
    #url = "http://172.16.3.155:8998/web/order/booking"
    for i in range(len(datas)):
        data1 = datas[i]
        data1['bookingOrderDTO']['aliOrderNo']="ALS145634"+str(random.randint(1000000,99999999))
        payload=json.dumps(data1)
        headers = {
          'Content-Type': 'application/json'
        }

        response = requests.request("POST", url, headers=headers, data=payload)
    return str(data1['bookingOrderDTO']['aliOrderNo'])+response.text
