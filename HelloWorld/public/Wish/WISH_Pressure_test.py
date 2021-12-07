import json
import random

import openpyxl
import requests
from locust import HttpUser,TaskSet,task


class Test(TaskSet):#偏远费
    def on_start(self):
        wb = openpyxl.load_workbook('../Wish/Wish_order.xlsx')
        ws = wb.active
        data = []
        for i in range(2,3000):#ws.max_row
            x=str(ws['A'+str(i)].value)
            data.append(json.loads(x))
        self.data = data

    @task(2)
    def Free_TH(self):#偏远费    下单
        header = {
            'Content-Type':'application/json',
            "Authorization":"Bearer"+" "+self.token
            }
        data1 = self.data[random.randint(1,len(self.data))]
        tracking_id = "WOSP10"+str(random.randint(1000000,99999999))
        data1["tracking_id"] =tracking_id
        data1["api_key"] = "r7ozJZmjvQHONMQZ9FIdCgjOxL02lWkaC55AGEQdJnPdlAqYhXbWRhPM6VUA"
        payload=json.dumps(data1)
        #下单
        with self.client.post('/tms-saas-web/wish/order/create', data = payload, headers = header, name = "WISH下单", catch_response = True) as response:
            if tracking_id in response.text:
                response.success()
            else:
                response.failure(response.text)
                failder_response = open('../request_data/failed_response.txt', 'ab')
                failder_response.write(str(response.text).encode('utf-8'))
                failder_response.close()
                datas = json.dumps(data1)
                failder_data = open('../request_data/failed_data.txt', 'ab')
                failder_data.write(str(datas+"\n").encode('utf-8'))
                failder_data.close()






class websitUser(HttpUser):
    tasks = [Test]
    host = "http://120.24.31.239:20000"
    min_wait = 1000  # 单位为毫秒
    max_wait = 2000  # 单位为毫秒
