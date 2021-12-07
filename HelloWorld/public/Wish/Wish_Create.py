import json
import os
import random
import openpyxl
import requests

def open_excel():
    wb = openpyxl.load_workbook(str(os.getcwd()+"/HelloWorld/public/Wish/Wish_order.xlsx"))
    ws = wb.active
    data = []
    for i in range(99,100):#ws.max_row
        x=str(ws['A'+str(4)].value)
        data.append(json.loads(x))
    return data

def Wish_Create():
    datas = open_excel()
    url = "http://120.24.31.239:20000/tms-saas-web/wish/order/create"
    # url = "http://120.79.131.69:20000/tms-saas-web/wish/order/create"#Kparcel生产环境测试地址
    logistics_order_code=[]
    for i in range(len(datas)):
        print(i+2)
        data1 = datas[i]
        data1["tracking_id"] ="WOSP0"+str(random.randint(1000000,99999999))
        data1["parcel"]["weight"] =4
        data1["api_key"] = "r7ozJZmjvQHONMQZ9FIdCgjOxL02lWkaC55AGEQdJnPdlAqYhXbWRhPM6VUA"
        payload=json.dumps(data1)
        headers = {
          'Content-Type': 'application/json'
        }
        response = requests.request("POST", url, headers=headers, data=payload)

        if json.loads(response.text)["message"] == "success":
            logistics_order_code.append({"logistics_order_code":json.loads(response.text)["logistics_order_code"]})
        else:
            print("下单"+response.text)
    return logistics_order_code
