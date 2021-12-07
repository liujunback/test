import json
import os
import random
import openpyxl

def open_excel():#读取ececl数据
    wb = openpyxl.load_workbook(str(os.getcwd()+"/HelloWorld/public/JD/JD_data.xlsx"))
    ws = wb.active
    data = []
    for i in range(2,5000):#ws.max_row
        x=str(ws['A'+str(i)].value)
        if x !="NULL":
            data.append(json.loads(x))
    return data


def JD_data():
    datas = open_excel()
    req = {
                "response":{
                    "content":{
                        "data":datas[0],
                        "exception":"",
                        "status":"SUCCESS"
                    },
                    "code":0
                }
            }
    return req
